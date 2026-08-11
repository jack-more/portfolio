#!/usr/bin/env python3
"""Generate public/sko-tracking.xlsx from app/SKOtracking/data.json.

The JSON is the single source of truth. The web page at /SKOtracking and this
workbook are both projections of it, so they cannot drift.

Run:  python3 scripts/build-tracking-xlsx.py
"""

import json
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA = ROOT / "app" / "SKOtracking" / "data.json"
OUT = ROOT / "public" / "sko-tracking.xlsx"

INK = "1C1917"
GREEN = "3A7D44"
AMBER = "8A6414"
RED = "8C2F21"
PARCHMENT = "F3EEE7"

HEAD_FONT = Font(bold=True, size=9, color="FFFFFF", name="Helvetica Neue")
BODY_FONT = Font(size=10, name="Helvetica Neue")
BOLD = Font(size=10, bold=True, name="Helvetica Neue")
HEAD_FILL = PatternFill("solid", fgColor=INK)
ZEBRA = PatternFill("solid", fgColor=PARCHMENT)
THIN = Side(style="thin", color="D8D0C6")
BORDER = Border(bottom=THIN)

HEALTH_COLOR = {
    "live": GREEN,
    "degraded": AMBER,
    "unverified": AMBER,
    "broken": RED,
}
SEV_COLOR = {"blocker": RED, "high": AMBER, "medium": "6B6560"}


def write_sheet(ws, headers, rows, widths, color_col=None, color_map=None):
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22

    for r, row in enumerate(rows, start=2):
        for i, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
            if r % 2 == 0:
                c.fill = ZEBRA
            if i == 1:
                c.font = BOLD
            if color_col and i == color_col and color_map:
                key = str(val).lower()
                if key in color_map:
                    c.font = Font(
                        size=10, bold=True, color=color_map[key], name="Helvetica Neue"
                    )

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    data = json.loads(DATA.read_text())
    wb = Workbook()

    # --- Sources ---
    ws = wb.active
    ws.title = "Sources"
    write_sheet(
        ws,
        ["Platform", "ID", "Loads live", "Browser events", "Server copy",
         "Dedupe key", "Secrets", "Hook point", "Health", "Note"],
        [
            [
                s["platform"],
                s["id"],
                "Yes" if s["runtime"] else "No",
                ", ".join(s["browserEvents"]),
                s["serverEvents"] or "—",
                s["dedupeKey"] or "—",
                ", ".join(s["secrets"]) or "—",
                s["hookPoint"],
                s["health"],
                s["note"],
            ]
            for s in data["sources"]
        ],
        [24, 26, 10, 34, 34, 24, 30, 34, 13, 62],
        color_col=9,
        color_map=HEALTH_COLOR,
    )

    # --- Weekly scorecard (the sheet that gets filled in each week) ---
    ws = wb.create_sheet("Weekly scorecard")
    ws.cell(row=1, column=1, value=f"Week of {data['weekOf']}").font = Font(
        bold=True, size=12, name="Helvetica Neue"
    )
    start = 3
    heads = ["Metric", "Last week", "Next week goal", "Where it comes from"]
    for i, h in enumerate(heads, start=1):
        c = ws.cell(row=start, column=i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
    for r, m in enumerate(data["scorecard"], start=start + 1):
        ws.cell(row=r, column=1, value=m["metric"]).font = BOLD
        for col in (2, 3):
            key = "actual" if col == 2 else "goal"
            c = ws.cell(row=r, column=col, value=m[key])
            c.font = BODY_FONT
            c.border = BORDER
            # Leave genuinely empty so a blank never reads as zero.
            if m["unit"] == "usd":
                c.number_format = '"$"#,##0;[Red]-"$"#,##0'
            elif m["unit"] == "x":
                c.number_format = '0.00"×"'
            else:
                c.number_format = "#,##0"
        n = ws.cell(row=r, column=4, value=m["source"])
        n.font = Font(size=9, color="6B6560", name="Helvetica Neue")
        n.alignment = Alignment(vertical="top", wrap_text=True)
    for i, w in enumerate([30, 16, 16, 70], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    note_row = start + len(data["scorecard"]) + 2
    ws.cell(row=note_row, column=1, value="On UGC attribution").font = BOLD
    n = ws.cell(row=note_row + 1, column=1, value=data["ugcNote"])
    n.font = Font(size=9, name="Helvetica Neue")
    n.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 3, end_column=4)

    # --- Top ads ---
    ws = wb.create_sheet("Top ads")
    write_sheet(
        ws,
        ["Ad", "Platform", "Spend", "Revenue", "ROAS", "Why it works"],
        [
            [a["ad"], a["platform"], a["spend"], a["revenue"], a["roas"], a["why"]]
            for a in data["topAds"]
        ]
        or [["—", "—", None, None, None, "No ad has delivered an impression yet."]],
        [34, 14, 12, 12, 10, 64],
    )

    # --- Coverage ---
    ws = wb.create_sheet("Funnel coverage")
    write_sheet(
        ws,
        ["Step", "OpenAI", "Meta", "TikTok"],
        [[c["step"], c["openai"], c["meta"], c["tiktok"]] for c in data["coverage"]],
        [22, 22, 22, 24],
    )

    # --- Issues ---
    ws = wb.create_sheet("Open issues")
    write_sheet(
        ws,
        ["ID", "Severity", "Platform", "Title", "Detail", "Fix", "Status"],
        [
            [i["id"], i["severity"], i["platform"], i["title"], i["detail"],
             i["fix"], i["status"]]
            for i in data["issues"]
        ],
        [8, 12, 18, 38, 68, 62, 34],
        color_col=2,
        color_map=SEV_COLOR,
    )

    # --- Campaign ---
    ws = wb.create_sheet("Campaign")
    labels = {
        "name": "Campaign", "account": "Ad account", "adSet": "Ad set",
        "objective": "Objective", "budget": "Budget", "ceiling": "Ceiling",
        "status": "Status", "ads": "Creative", "promo": "Promo",
    }
    write_sheet(
        ws,
        ["Field", "Value"],
        [[labels.get(k, k), v] for k, v in data["campaign"].items()],
        [18, 58],
    )
    last = ws.max_row + 2
    ws.cell(row=last, column=1, value="Verified").font = BOLD
    ws.cell(row=last, column=2, value=data["lastVerified"]).font = BODY_FONT

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(ROOT)}  ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
